import tkinter
import sqlite3
import datetime
import pywin
import pywin32_system32
import pywintypes
import win32print
import openpyxl
from win32api import ShellExecute
from tkinter import *
from tkinter import ttk,messagebox,filedialog


class aplicativo():
    def __init__(self):
        self.window = Tk()
        self.window.title("FLProgram")
        self.window.geometry("800x600")
        self.window.state("zoomed")
        self.window["bg"] = "red"
        self.icone = "c:\FLPrograms\Read\Config\Icone.ico"
        self.diretorios()
        a = self.dirDBgastos[-4:].lower()
        b = self.dirDBfechamento[-4:].lower()
        if a != ".sql" or b != ".sql":
            self.window.destroy()
            erro = messagebox.askokcancel(title="Falha de Banco de Dados", message="ENCONTRADO ERRO NO BANCO DE DADOS! \n CHAMAR ASSISTENCIA TÉCNICA.")
            atcnica = messagebox.askyesno(title="ERRO", message="Confirmar ir à area de restauração")
            if atcnica:
                ast = Tk()
                ast.title("CONFIRM USER")
                ast.geometry("300x200")
                ast.resizable(0,0)
                name = Label(ast, text="User", font="arial 12", bg="gray50", width=8)
                ename = Entry(ast, font="arial 12", width=15, show="&")
                name.place(x=40 , y=60)
                ename.place(x=130 , y=60)
                chave = Label(ast, text="Password", font="arial 12", bg="gray50", width=8)
                echave = Entry(ast, font="arial 12", width=15, show="&")
                chave.place(x=40 , y=85)
                echave.place(x=130 , y=85)
                def confirmar_usuario():
                    user = ename.get().lower()
                    passw = echave.get()
                    if user == "felipe" and passw == "FLProgram":
                        self.cadastrarservidor()
                        ast.destroy()
                    else:
                        lblerro = Label(ast, text="ACESS DENIED", font="arial 12", bg="black", fg="red")
                        lblerro.place(x=80, y=150)
                botn = Button(ast, text="Login", font="arial 12", width=8, command=confirmar_usuario)
                botn.place(x=130 , y=115)
                
                ast.mainloop()
        self.DB()
        self.fontebotao = "arial 15 bold"
        self.fontetitulo = "times 20 bold"
        self.fonteLabel = "arial 12 bold"

        self.titulo = Label(self.window, text="SUPERMERCADO HELY", bg="red" , font=self.fontetitulo)
        self.titulo.pack(anchor=CENTER, pady=30)
        
        self.btnfechamento = Button(self.window,text="Fechamento de caixa", font=self.fontebotao , width=23, bg="yellow",command=self.fechamento)
        self.btnfechamento.pack(anchor=CENTER, pady=2)
        self.btnfechamento.focus()
        self.btncont = Button(self.window,text="Contar dinheiro",bg="yellow", width=23, height=1, font=self.fontebotao , command=self.contarDinheiro)
        self.btncont.pack(anchor=CENTER, pady=2)
        self.btnfechar = Button(self.window,text="Fechar",bg="yellow", width=23, height=1 , font=self.fontebotao ,command=self.window.destroy)
        self.btnfechar.pack(anchor=CENTER, pady=2)
        self.btnadmin = Button(self.window,text="Administrativo", width=23, height=1,font="arial 11 bold", background="#808080" , command=self.confirmar)
        self.btnadmin.pack(anchor=CENTER, pady=30)

        self.window.iconbitmap(self.icone)
        self.window.mainloop()
        
    def voltar(self):
        try:
            self.window1.destroy()
        except:
            pass
        try:
            self.window2.destroy()
        except:
            pass
        self.back = aplicativo()

    def diretorios(self):
        
        with open(r"c:\FLPrograms\Read\Config\Texts\Dbg.txt" ,"r" ) as dbg:
            self.dirDBgastos = dbg.read()
        with open(r"c:\FLPrograms\Read\Config\Texts\Dbf.txt" ,"r" ) as dbf:
            self.dirDBfechamento = dbf.read()
        with open(r"c:\FLPrograms\Read\Config\Texts\Impressora padrao.txt" ,"r" ) as printer:
            self.impressoraPadrao = printer.read()
        with open(r"c:\FLPrograms\Read\Config\Texts\Users.txt" ,"r" ) as users:
            self.lastuser = users.read()
        
        todas_impressoras = win32print.EnumPrinters(2)
        self.lista_impressoras = []
        for row in todas_impressoras:
            impressora = row[2]
            self.lista_impressoras.append(impressora)
            
    def DB(self):
        # gastos -> tabela gastos / idcontrole integer primary key autoincrement , dia text , gastosSemRetorno text , valorSR float , gastosComRetorno text , valorCR float , funcionario text
        # fechamento -> tabela fechamento / idcontrole integer primary key autoincrement ,dia text ,caixa text,  dinheiro float, cartao float, aprazo float , pix float , gastos float , supervisor text,senha text, impressora text
        self.bankgastos = sqlite3.connect(self.dirDBgastos)
        self.bankfechamento = sqlite3.connect(self.dirDBfechamento)
        self.cursorgastos = self.bankgastos.cursor()
        self.cursorfechamento = self.bankfechamento.cursor()
        self.cursorfechamento.execute("CREATE TABLE IF NOT EXISTS fechamento(idcontrole integer primary key autoincrement ,dia text ,caixa text,  dinheiro float, cartao float, aprazo float , pix float , gastos float , supervisor text,senha text, impressora text) ")
        self.cursorgastos.execute("CREATE TABLE IF NOT EXISTS gastos(idcontrole integer primary key autoincrement ,dia text , gastosSemRetorno text ,valorSR float , gastosComRetorno text , valorCR float , funcionario text) ")
        self.bankfechamento.commit()
        self.bankgastos.commit()
    
    def fechamento(self):
        self.window1 = Tk()
        try:
            self.window.destroy()
        except:
            pass
        self.window1.geometry("800x600")
        self.window1.state("zoomed")
        self.window1["bg"] = "red"
            
        self.btnvoltar = Button(self.window1, text="Voltar", command=self.voltar)
        self.btnvoltar.place(x=2 , y=2)
        
        
        self.titulo = Label(self.window1,text="FECHAMENTO DE CAIXA" , bg="red" , font=self.fontetitulo)
        self.titulo.place(x=250, y=20)
        
        self.lblcx = Label(self.window1,text="CX", width=3 , relief="solid" , bd=1 ,  bg="yellow" , font=self.fonteLabel)
        self.lblcx.place(x=40, y=100)
        self.lbcx = Entry(self.window1,width=10 , font=self.fonteLabel)
        self.lbcx.place(x=76, y=100)
        self.lbcx.insert( 0 , "N°/Nome")
        self.lbcx.focus()
        
        self.ldia = Label(self.window1,text="Dia", width=3, bd = 1, relief = "solid", font=self.fonteLabel, bg="#ffff00")
        self.ldia.place(x=215 , y=100)
        self.txtdia = Entry(self.window1, font=self.fonteLabel , width=10)
        self.txtdia.place(x=250 , y=100)
        
        
        
        self.datual = datetime.datetime.now()
        self.mesatual = self.datual.month
        self.diaatual = self.datual.day
        
        if self.mesatual < 10 and self.diaatual < 10 :
            self.txtdia.insert(0,'0' + str(self.datual.day) + "/" + '0' + str(self.datual.month) + "/" + str(self.datual.year) )
        elif self.mesatual < 10:
            self.txtdia.insert(0, str(self.diaatual) + "/" + '0' + str(self.mesatual) + "/" + str(self.datual.year) )
        elif self.diaatual < 10:
            self.txtdia.insert(0,'0' + str(self.datual.day) + "/" + str(self.datual.month) + "/" + str(self.datual.year) )
        else:
            self.txtdia.insert(0, str(self.datual.day) + "/" + str(self.datual.month) + "/" + str(self.datual.year) )
        
            
        self.dmanha = Label(self.window1,text="Dinheiro da Manhã", width=20 , anchor=W , font=self.fonteLabel,bd = 1, relief = "solid", bg="#ffff00")
        self.dmanha.place(x=40,y=130 )
        self.txtmanha = Entry(self.window1,width=10 , font=self.fonteLabel)
        self.txtmanha.place(x=250 , y=130)
        
        self.dtarde = Label(self.window1,text="Dinheiro da Tarde", anchor=W , font=self.fonteLabel, bd = 1, relief = "solid",width=20,bg="#ffff00")
        self.dtarde.place(x=40 , y=155)
        
        with open(r"c:\FLPrograms\Read\Config\Texts\Tarde.txt",'r') as arquivo:
            for row in arquivo:
                self.valortarde = float(row)
                
        with open(r"c:\FLPrograms\Read\Config\Texts\Tarde.txt","w") as arq:
            arq.write(str("0"))
            arq.close()
            
        if self.valortarde > 0 :
            self.txtarde1 = Label(self.window1,text=f"{self.valortarde}", bd=1, anchor=E ,width=9, font=self.fonteLabel, bg="white")
            self.txtarde1.place(x=250 , y=155)
        else:
            self.txttarde = Entry(self.window1,width=10  , font=self.fonteLabel)
            self.txttarde.place(x=250 , y= 155)

            
        self.supri = Label(self.window1, text="Suprimento",width=20, anchor=W , font=self.fonteLabel ,bd = 1, relief = "solid", bg="#ffff00")
        self.supri.place(x=40 , y=180)
        self.txtsupri = Entry(self.window1,width=10 , font=self.fonteLabel)
        self.txtsupri.place(x=250 , y=180)
        
        
        self.card = Label(self.window1 , text="Cartão" , width=20 , anchor=W , font=self.fonteLabel , bd = 1, relief = "solid", bg="#ffff00")
        self.card.place(x=40 , y=205)
        self.txtcard = Entry(self.window1, width=10 , font=self.fonteLabel)
        self.txtcard.place(x=250 , y=205)
        
        self.apra = Label(self.window1,text="A Prazo",width=20 , anchor=W , font=self.fonteLabel , bd = 1, relief = "solid", bg="#ffff00")
        self.apra.place(x=40 , y=230)
        self.txtapra = Entry(self.window1 , width=10 , font=self.fonteLabel)
        self.txtapra.place(x=250 , y=230)
        
        self.lpix = Label(self.window1 , text="Pix/ticket", anchor=W , width=20, bd = 1, relief = "solid",font=self.fonteLabel , bg="#ffff00")
        self.lpix.place(x=40 , y=255)
        self.txtpix = Entry(self.window1 , width=10 , font=self.fonteLabel)
        self.txtpix.place(x=250 , y=255)
        
        self.sistem = Label(self.window1,text="Saldo Total",width=20, anchor=W , bd = 1, relief = "solid", font=self.fonteLabel, bg="#ffff00")
        self.sistem.place(x=40 , y=280)
        self.txtsistem = Entry(self.window1 , width=10 , font=self.fonteLabel)
        self.txtsistem.place(x=250 , y=280)
        self.botao = Button(self.window1,text="Confirmar" , width=8 ,bg="yellow", height=1 , font=self.fontebotao , command=self.calculo )
        self.botao.place(x=250 , y=305)
        self.window1.iconbitmap(self.icone)
        self.window1.mainloop()
         
    def calculo(self):
        
        self.dia = self.txtdia.get()
        
        if len(self.txtmanha.get()) > 0 :
            self.dinheiromanha = float(self.txtmanha.get().replace(",","."))
        else:
            self.dinheiromanha = 0
        if self.valortarde > 0:
            self.dinheirotarde = self.valortarde
        elif len(self.txttarde.get()) > 0:
            self.dinheirotarde = float(self.txttarde.get().replace(",","."))
        else:
            self.dinheirotarde = 0
        if len(self.txtsupri.get()) > 0 :
            self.suprimento = float(self.txtsupri.get().replace(",","."))
        else:
            self.suprimento = 0
        self.dinheirocsuprimento = self.dinheiromanha + self.dinheirotarde
        self.dinheirototal = self.dinheirocsuprimento - self.suprimento
        if len(self.txtcard.get()) > 0 :
            self.cartao = float(self.txtcard.get().replace(",","."))
        else:
            self.cartao = 0
        if len(self.txtapra.get()) > 0 :
            self.aprazo = float(self.txtapra.get().replace(",","."))
        else:
            self.aprazo = 0
        if len(self.txtpix.get()) > 0 :
            self.pix = float(self.txtpix.get().replace(",","."))
        else:
            self.pix = 0
        if len(self.txtsistem.get()) > 0 :
            self.sistema = float(self.txtsistem.get().replace(",","."))
        else:
            self.sistema = 0
        
        self.cx = self.lbcx.get().upper()
        self.somatotal = (self.dinheirototal + self.cartao + self.aprazo + self.pix )
        self.saldof = self.somatotal - self.sistema
        self.saldofinal = format(self.saldof , '.2f')
        
        self.lin1 = Label(self.window1, text=f"CX {self.cx}" , width=15, font=self.fonteLabel)
        self.lin1.place(x=400 , y=100)
        
        self.lin2 = Label(self.window1,text="Dinheiro da Manhã",font=self.fonteLabel,bg = "#ffff00", width=20, bd=1, relief='solid',anchor=W )
        self.lin2.place(x=400 , y=130)
        self.lin3 = Label(self.window1,text="Dinheiro da Tarde",font=self.fonteLabel, width=20,bg = "#ffff00", bd=1, relief='solid',anchor=W )
        self.lin3.place(x=400 , y=155)
        self.lin4 = Label(self.window1,text="Suprimento",bg = "#ffff00",font=self.fonteLabel, width=20, bd=1, relief='solid',anchor=W )
        self.lin4.place(x=400 , y=205)
        self.lin5 = Label(self.window1,text="Cartão",font=self.fonteLabel, width=20,bg = "#ffff00", bd=1, relief='solid',anchor=W )
        self.lin5.place(x=400 , y=255)
        self.lin6 = Label(self.window1,text="A Prazo",font=self.fonteLabel, width=20, bd=1,bg = "#ffff00", relief='solid',anchor=W )
        self.lin6.place(x=400 , y=280)
        self.lin7 = Label(self.window1,text="Pix ou Ticket",font=self.fonteLabel, width=20, bd=1,bg = "#ffff00", relief='solid',anchor=W )
        self.lin7.place(x=400 , y=305)
        self.lin9 = Label(self.window1,text="Saldo Total",font=self.fonteLabel, width=20, bd=1,bg = "#ffff00", relief='solid',anchor=W )
        self.lin9.place(x=400 , y=330)
        self.lin8 = Label(self.window1,text="Sistema",font=self.fonteLabel, width=20, bd=1, relief='solid',bg = "#ffff00",anchor=W )
        self.lin8.place(x=400 , y=355)
        self.lin9 = Label(self.window1,text="Saldo Final",font=self.fonteLabel, width=20, bd=1,bg = "#ffff00", relief='solid',anchor=W )
        self.lin9.place(x=400 , y=380)
        
        self.lin22 = Label(self.window1,text=f" {format(self.dinheiromanha , '.2f') }",font=self.fonteLabel, width=10, bd=1, relief='solid',bg = "#ffff00",anchor=E )
        self.lin22.place(x=610 , y=130)
        self.lin33 = Label(self.window1,text=f"+  {format(self.dinheirotarde , '.2f') }",font=self.fonteLabel, width=10, bd=1, relief='solid',bg = "#ffff00",anchor=E )
        self.lin33.place(x=610 , y=155)
        self.lin44 = Label(self.window1,text=f"{format(self.dinheirocsuprimento , '.2f') }",font=self.fonteLabel, width=10, bd=1, relief='solid',anchor=E )
        self.lin44.place(x=610 , y=180)
        self.lin55 = Label(self.window1,text=f"-  {format(self.suprimento , '.2f') }",font=self.fonteLabel, width=10, bd=1, relief='solid',bg = "#ffff00",anchor=E )
        self.lin55.place(x=610 , y=205)
        self.lin66 = Label(self.window1,text=f"{format(self.dinheirototal , '.2f') }",font=self.fonteLabel, width=10, bd=1, relief='solid',anchor=E )
        self.lin66.place(x=610 , y=230)
        self.lin77 = Label(self.window1,text=f"{format(self.cartao , '.2f' ) }",font=self.fonteLabel, width=10,bg = "#ffff00", bd=1, relief='solid',anchor=E )
        self.lin77.place(x=610 , y=255)
        self.lin88 = Label(self.window1,text=f"{format(self.aprazo , '.2f' ) }",font=self.fonteLabel, width=10, bd=1,bg = "#ffff00", relief='solid',anchor=E )
        self.lin88.place(x=610 , y=280)
        self.lin99 = Label(self.window1,text=f"+  {format(self.pix , '.2f' ) }",font=self.fonteLabel, width=10, bd=1,bg = "#ffff00", relief='solid',anchor=E )
        self.lin99.place(x=610 , y=305)
        self.lin91 = Label(self.window1,text=f"{format(self.somatotal , '.2f') }",font=self.fonteLabel, width=10, bd=1, relief='solid',anchor=E )
        self.lin91.place(x=610 , y=330)
        self.lin81 = Label(self.window1,text=f"-  {format(self.sistema , '.2f') }",font=self.fonteLabel, width=10, bd=1,bg = "#ffff00", relief='solid',anchor=E )
        self.lin81.place(x=610 , y=355)
        
        if self.somatotal > self.sistema :
            self.lin71 = Label(self.window1,text=f"{self.saldofinal }",font=self.fonteLabel, width=10, bd=2,bg = "#00ff00", relief='solid',anchor=E )
            self.lin71.place(x=610 , y=380) 
        elif self.somatotal == self.sistema:
            self.lin71 = Label(self.window1,text=f"{self.saldofinal }",font=self.fonteLabel, width=10, bd=2, bg="white" , relief='solid',anchor=E )
            self.lin71.place(x=610 , y=380)
        else:
            self.lin71 = Label(self.window1,text=f"{self.saldofinal }",font=self.fonteLabel, width=10, bd=2,bg ="#ff9900" , fg="red1" , relief='solid',anchor=E )
            self.lin71.place(x=610 , y=380)
        
        self.btnimprimir = Button(self.window1,text="Imprimir",background="yellow" ,font=self.fontebotao, width=8 , command=self.imprimir)
        self.btnimprimir.place(x=460 , y=415)
        
        with open(r'C:\FLPrograms\Read\Config\Texts\Fechamento impresso.txt','w') as arquivo:
                arquivo.write(str(f"""
-------------------------------------
-------------------------------------
         SUPERMERCADO HELY
        FECHAMENTO DE CAIXA

CAIXA: {self.cx}     DATA:{self.dia}
_____________________________________
MANHA                       {self.dinheiromanha}
TARDE                       {self.dinheirotarde}
______________________________________
TOTAL                       {self.dinheirocsuprimento}
SUPRIMENTO                - {self.suprimento}
______________________________________
TOTAL                       {self.dinheirototal}
CARTÃO                      {self.cartao}
A PRAZO                     {self.aprazo}
PIX / TICKET                {self.pix}
______________________________________
TOTAL                       {format(self.somatotal, '.2f')}
SISTEMA                     {self.sistema}
______________________________________
SALDO                       {self.saldofinal}



--------------------------------------
--------------------------------------

"""))
                
    
    def imprimir(self):
        try:
            self.enviar = Button(self.window1,text="Enviar", font=self.fontebotao , bg="yellow", width=8 , command=self.guardar)
            self.enviar.place(x=580 , y=415)
            win32print.SetDefaultPrinter(self.impressoraPadrao)
            caminho = r"c:\FLPrograms\Read\Config\Texts"
            ShellExecute(0, "print" , "Fechamento impresso.txt" ,  None , caminho , 0  )
        except:
            messagebox.askokcancel(title="ERRO", message="Falha ao imprimir arquivo! \nContatar suporte tecnico" 

    def guardar(self):
        
        confirm = messagebox.askyesno(title="Confirmar Envio", message="Confirmar e Fechar Fechamento de caixa")
        
        if confirm:
            
            dia = self.dia
            caixa = self.cx
            dinheiro = self.dinheirototal
            cartao = self.cartao
            aprazo = self.aprazo
            pix = self.pix
            self.cursorfechamento.execute(f"INSERT INTO fechamento (dia,caixa, dinheiro, cartao, aprazo, pix) VALUES ('{dia}','{caixa}',{dinheiro},{cartao},{aprazo},{pix} )")
            self.bankfechamento.commit()
            self.window1.destroy()

    def contarDinheiro(self):
        
        self.window2 = Tk()
        self.window.destroy()
        self.window2.title("Contar Dinheiro")
        self.window2.geometry("800x600")
        self.window2.state("zoomed")
        self.window2["bg"] = "red"
        
        self.btnback = Button(self.window2, text="voltar", command=self.voltar)
        self.btnback.place(x=2 , y=2)
        
        self.lbl = Label(self.window2,text="CONTAGEM DE DINHEIRO",font=self.fontetitulo , bg="red" ,  width=25)
        self.lbl.place(x=250 , y=30)
        
        self.lbl1 = Label(self.window2,text="Obs: Preencher com os valores totais somados de cada nota",font=self.fonteLabel , anchor=W,bd=1,relief="solid")
        self.lbl1.place(x=70 , y=550)
        self.lbl2 = Label(self.window2,text="Notas 200/100/50",font=self.fonteLabel, width=15,bg= "#ffff00",bd=1,anchor=W,relief="solid")
        self.lbl2.place(x=40 , y=175)
        self.lbl3 = Label(self.window2,text="Notas de 20,00",bg= "#ffff00",font=self.fonteLabel, width=15,bd=1,anchor=W,relief="solid")
        self.lbl3.place(x=40 , y=200)
        self.lbl4 = Label(self.window2,text="Notas de 10,00",font=self.fonteLabel ,bg= "#ffff00", width=15,bd=1,anchor=W,relief="solid")
        self.lbl4.place(x=40 , y=225)
        self.lbl5 = Label(self.window2,text="notas de 5,00",font=self.fonteLabel , width=15,bd=1,bg= "#ffff00",anchor=W,relief="solid")
        self.lbl5.place(x=40 , y=250)
        self.lbl6 = Label(self.window2,text="Notas de 2,00",bg= "#ffff00",font=self.fonteLabel , width=15,bd=1,anchor=W,relief="solid")
        self.lbl6.place(x=40 , y=275)
        self.lbl7 = Label(self.window2,text="Moedas",font=self.fonteLabel ,bg= "#ffff00", width=15,bd=1,anchor=W,relief="solid")
        self.lbl7.place(x=40 , y=300)
        self.lbl8 = Label(self.window2,text="Total",font=self.fonteLabel , width=15,bd=1, bg= "#ffff00",anchor=W,relief="solid")
        self.lbl8.place(x=40 , y=325)
        
        self.tlbl2 = Entry(self.window2,width=10,font=self.fonteLabel)
        self.tlbl3 = Entry(self.window2,width=10,font=self.fonteLabel)
        self.tlbl4 = Entry(self.window2,width=10,font=self.fonteLabel)
        self.tlbl5 = Entry(self.window2,width=10,font=self.fonteLabel)
        self.tlbl6 = Entry(self.window2,width=10,font=self.fonteLabel)
        self.tlbl7 = Entry(self.window2,width=10,font=self.fonteLabel)
        self.tlbl2.focus()
        
        self.tlbl2.place(x=200 , y=175)
        self.tlbl3.place(x=200 , y=200)
        self.tlbl4.place(x=200 , y=225)
        self.tlbl5.place(x=200 , y=250)
        self.tlbl6.place(x=200 , y=275)
        self.tlbl7.place(x=200 , y=300)

        self.tlbl8 = Label(self.window2,text="0,00", width=9 , bg="#ffff00", anchor=E ,font=self.fonteLabel)
        self.tlbl8.place(x=200 , y=325)
    
        
        self.btn = Button(self.window2,text="Confirmar" ,width=10 , font="arial 12" , bg="#4dc3ff" , command=self.somar)
        self.btn.place(x=100 , y= 355)
        self.window2.iconbitmap(self.icone)
        self.window2.mainloop()
    
    def somar(self):
        
        if len(self.tlbl2.get()) > 0:
            v2 = float(self.tlbl2.get().replace(",","."))
        else:
            v2 = 0
        if len(self.tlbl3.get()) > 0:
            v3 = float(self.tlbl3.get().replace(",","."))
        else:
            v3 = 0
        if len(self.tlbl4.get()) > 0:
            v4 = float(self.tlbl4.get().replace(",","."))
        else:
            v4 = 0
        if len(self.tlbl5.get()) > 0:
            v5 = float(self.tlbl5.get().replace(",","."))
        else:
            v5 = 0
        if len(self.tlbl6.get()) > 0:
            v6 = float(self.tlbl6.get().replace(",","."))
        else:
            v6 = 0
        if len(self.tlbl7.get()) > 0:
            v7 = float(self.tlbl7.get().replace(",","."))
        else :
            v7 = 0
        
        self.vt = (v2+v3+v4+v5+v6+v7)
        self.tlbl8["text"] = f"{format(self.vt , '.2f')}"
        
    
        def importar():
            with open(r"c:\FLPrograms\Read\Config\Texts\Tarde.txt","w") as trd:
                trd.write(str(self.vt))
            self.window2.destroy()
            self.fechamento()

        self.btn1 = Button(self.window2, text="Fechamento", font="arial 12" , width=10, bg="#4dc3ff", command=importar)
        self.btn1.place(x=200 , y=355)

    def confirmar(self):
        
        self.jnl1 = Tk()
        self.jnl1.geometry("300x300")
        self.jnl1.title("Logon")
        self.jnl1.resizable(0,0)
        self.nome = Label(self.jnl1, text="Usuario",font= "arial 13",width=10)
        self.nome.pack(padx=0 , pady=5,  anchor=CENTER)
        self.tnome = Entry(self.jnl1, width=10, background="#ffffff", font="arial 13")
        self.tnome.pack(padx=5, pady=0, anchor=CENTER)
        self.tnome.focus()
        self.Key = Label(self.jnl1, text="Senha", font="arial 13")
        self.Key.pack(padx=0, pady=5, anchor=CENTER)
        self.ekey = Entry(self.jnl1, width=10,show="*",font="arial 13",background="#ffffff")
        self.ekey.pack(padx=5, pady=0, anchor=CENTER)
        
        self.btnconfirmar = Button(self.jnl1,text="Logar",width=5,font="arial 13", bg="gray70", command=self.verificar )
        self.btnconfirmar.place(x=90 , y=140)
        self.btnfechar = Button(self.jnl1,text="Fechar",width=5, font="arial 13", bg="gray70" , command=self.jnl1.destroy)
        self.btnfechar.place(x=150 , y=140)
        self.erro = Label(self.jnl1,text="", font="arial 13 bold" , width=15)
        self.erro.place(x=70 , y=200)
        self.lblobs = Label(self.jnl1, text='Para informacões de software logar com "INFO". ', font = "arial 10" )
        self.lblobs.place(x=5 , y=230)                
        self.jnl1.iconbitmap(self.icone)
        self.jnl1.mainloop()
    
    def verificar(self):
            users = self.tnome.get()
            user = users.lower()
            passwd = self.ekey.get()
            
            with open(r'c:\FLPrograms\Read\Config\Texts\Users.txt','w') as usuario:
                usuario.write(str(f"{user}"))
            
            self.cursorfechamento.execute("SELECT supervisor FROM fechamento")
            lista = self.cursorfechamento.fetchall()
            
            controle = 1
            controle1 = len(lista)
            
            for row in lista:
                
                if user == "felipe " and passwd =="FLProgram":
                    
                    self.jnl1.destroy()
                    self.Programador()
                    self.administrador()
                    break
                
                elif row == (f'{user}',):
                    b = lista.index((f'{user}',))
                
                    self.cursorfechamento.execute("SELECT senha FROM fechamento")
                    lista2 = self.cursorfechamento.fetchall()
                    
                    if lista2[b] == ((f'{passwd}',)):
                        
                        self.jnl1.destroy()
                        self.administrador()
                    
                    else:
                        self.erro["text"] = "ACESSO NEGADO !"
                        self.erro["bg"] = "#000"
                        self.erro["fg"] = "red"
                        break

                    
                
                elif user == "info":
                    
                    jnl2 = Tk()
                    jnl2.geometry("400x300")
                    jnl2.title("INFORMAÇÃO DE SOFTWARE")
                    
                    criador = Label(jnl2,text="FLPrograms", font="Times 12 bold", width=20)
                    criador.pack(anchor=CENTER)
                    versao = Label(jnl2,text="Vesão 1.0/03.2022", font="Times 12 bold", width=20)
                    versao.pack(anchor=CENTER)
                    cntt = Label(jnl2,text="Contato: felipesrodrigs@gmail.com \n (97)984165908", font="Times 12 bold", width=30)
                    cntt.pack(anchor=CENTER)
                    btnfechar = Button(jnl2,text="Fechar", font="Times 12 bold" , width=10, bg="#ffffff", command=jnl2.destroy)
                    btnfechar.pack(anchor=CENTER)
                
                    jnl2.mainloop                        
                    break
                elif controle == controle1:    
                    
                    self.erro["text"] = "ACESSO NEGADO !"
                    self.erro["bg"] = "#000"
                    self.erro["fg"] = "red"        
                    break

                else:
                    
                    controle += 1
    
    def Programador(self):
        self.jnl3 = Tk()
        self.jnl3.title("Configurações")    
        self.jnl3.geometry("800x600")
        self.jnl3.state("zoomed")
        
        btnprint = Button(self.jnl3,text="Selecionar impressora padrão", font=self.fontebotao , bg="#4dc3ff", width=25, command=self.print_padrao)
        btnprint.pack(anchor = CENTER, pady=60)
        
        btnsupervisor = Button(self.jnl3, text="Cadastrar novo Supervisor", font = self.fontebotao ,bg="#4dc3ff", width=25, command=self.new_super)
        btnsupervisor.pack(anchor=CENTER, pady=2)
        
        btndb = Button(self.jnl3, text="Cadastrar Servidor Local" , font=self.fontebotao , bg="#4dc3ff",width=25 , command=self.cadastrarservidor)
        btndb.pack(anchor=CENTER , pady=2)
        self.jnl3.iconbitmap(self.icone)
        self.jnl3.mainloop()
        
    def print_padrao(self):
        jnl4 = Tk()
        lblselect = Label(jnl4,text="Selecione a Impressora padrão", font="Arial 12", anchor=CENTER, width=30)
        lblselect.pack(pady=5)
        cb_print = ttk.Combobox(jnl4,values=self.lista_impressoras, font="arial 12",width=20)
        cb_print.pack(anchor=CENTER)
        cb_print.set(self.impressoraPadrao)
        def select():
            
            with open(r'c:\FLPrograms\Read\Config\Texts\Impressora padrao.txt','w') as impressorapadrao:
                impressorapadrao.write(cb_print.get())
                jnl4.destroy()
                
        btnconfirm = Button(jnl4,text="Selecionar e Fechar",width=20,bg="#4dc3ff", command=select )
        btnconfirm.pack(anchor=CENTER)
        jnl4.iconbitmap(self.icone)
        jnl4.mainloop()

    def new_super(self):
        jnl6 = Tk()
        jnl6.title("Casdastrar")
        jnl6.geometry("500x250")
        
        
        
        
        lblnome = Label(jnl6,text="Nome \n \nSenha", width=10, height=3 , font="arial 12")
        lblnome.place(x=20, y=20)
        
        enome = Entry(jnl6,width=20,font="arial 12")
        enome.place(x=100, y=20)
        
        esenha = Entry(jnl6,width=20, font="arial 12")
        esenha.place(x=100, y=60)
        
        
        def cadastrar():
            nome = enome.get().lower()
            senha = esenha.get()
            
            self.cursorfechamento.execute("SELECT supervisor FROM fechamento")
            supervisores = self.cursorfechamento.fetchall()
            control = 1
            control1 = len(supervisores)
            
            for nomes in supervisores:
                if nomes[0] == nome:
                    
                    repetido = messagebox.askyesno(title="Usuario Repetido!", message="Supervisor já cadastrado, \n deseja atualizar nova senha?")
                    if repetido:
                        self.cursorfechamento.execute(f"UPDATE fechamento SET senha = '{senha}' WHERE supervisor = '{nome}'")
                        self.bankfechamento.commit()
                        jnl6.destroy()
                        break
                    else:
                        pass
                elif control == control1:
                
                    confirma = messagebox.askyesno(title="", message="Confirmar Inclusao de novo Supervisor ?")
                    if confirma:
                        
                        self.cursorfechamento.execute(f"INSERT INTO fechamento (supervisor , senha) VALUES ('{nome}','{senha}')")
                        self.bankfechamento.commit()
                        jnl6.destroy() 
                        break
                else:
                    control +=1
            
        def apagar():
            nome = enome.get().lower()
            
            self.cursorfechamento.execute("SELECT supervisor FROM fechamento")
            supervir = self.cursorfechamento.fetchall()
            control = 1
            control1 = len(supervir)
            for name in supervir:
                if name[0] == nome:
                    
                    askdelete = messagebox.askyesno(title="!", message="APAGAR USUARIO ?")
                    if askdelete:
                        self.cursorfechamento.execute(f"DELETE FROM fechamento WHERE supervisor = '{nome}' ")
                        self.bankfechamento.commit()
                        jnl6.destroy()
                        break
                elif control == control1 :
                    messagebox.showinfo(title="", message="Usuario não encontrado")
                else:
                    control += 1
            
        btncadastrar = Button(jnl6, text="Cadastrar Novo supervisor", font ="arial 12", width=22,bg="#4dc3ff", command=cadastrar)
        btncadastrar.place(x=50, y= 120)
        
        btndelete = Button(jnl6,text="Apagar Supervisor", font= "arial 12", width=22, bg="red", command=apagar)
        btndelete.place(x=260, y= 120)
        jnl6.iconbitmap(self.icone)
        jnl6.mainloop()

    def cadastrarservidor(self):
        
        alterar = messagebox.askyesno(title="Alterar Diretorios", message="Selecionar Local dos Bancos de dados \n Fechamento \ Gastos")
        if alterar:
            localdbf = filedialog.askopenfilenames()[0]
            if localdbf != None:
                with open(r"c:\FLPrograms\Read\Config\Texts\Dbf.txt" , "w") as dbf:
                    dbf.write(localdbf)
            localdbg = filedialog.askopenfilenames()[0]
            if localdbg != None:    
                with open(r"c:\FLPrograms\Read\Config\Texts\Dbg.txt" , "w") as dbg:
                    dbg.write(localdbg)

    def administrador(self):
            
        jnl2 = Tk()
        jnl2.title("Administrativo")
        jnl2.geometry("800x600")
        jnl2.state("zoomed")
        jnl2["bg"] = "red"
        titulo = Label(jnl2,text="ADIMINISTRATIVO", width=20, bg="red", font=self.fontetitulo)
        titulo.pack(anchor=CENTER, pady=20)
        
        def gastos():
            jnl2.destroy()
            self.gerirgastos()
        
        btngast = Button(jnl2,text="Gerenciar Gastos", bg="#ff0", width=20 , font=self.fonteLabel, command=gastos)
        btngast.pack(anchor=CENTER,padx=20,pady=9)
        
        btnrelatorio = Button(jnl2,text="Consultar Relatorios", bg="#ff0",width=20, font=self.fonteLabel, command=self.relatorio)
        btnrelatorio.pack(anchor=CENTER,padx=20)
        jnl2.iconbitmap(self.icone)
        jnl2.mainloop()

    def gerirgastos(self):
        
        jnl5 = Tk()
        jnl5.title("Adicionar Gastos")
        jnl5.geometry("8000x600")
        jnl5.state("zoomed")
        
        def voltar():
            jnl5.destroy()
            self.administrador()
        
        btnvoltar = Button(jnl5,text="Voltar", font="arial 10", command=voltar)
        btnvoltar.place(x=2,y=2)
        
        titulo = Label(jnl5,text="GERENCIAR GASTOS", width=20, font=self.fontetitulo )
        titulo.place(x=250, y=40)
        
        lblday = Label(jnl5,text="Dia", font="arial 12 bold", bg="#fff" , bd=5 , width=4, anchor=W)
        lblday.place(x=50, y=40)
        edia = Entry(jnl5,width=10, font="arial 12")
        edia.place(x=50,y=70)
        
        datual = datetime.datetime.now()
        mesatual = datual.month
        diaatual = datual.day
        if mesatual < 10 and diaatual < 10 :
            edia.insert(0,'0' + str(datual.day) + "/" + '0' + str(datual.month) + "/" + str(datual.year) )
        elif mesatual < 10:
            edia.insert(0, str(datual.day) + "/" + '0' + str(datual.month) + "/" + str(datual.year) )
        elif diaatual < 10:
            edia.insert(0,'0' + str(datual.day) + "/" + str(datual.month) + "/" + str(datual.year) )
        else:
            edia.insert(0, str(datual.day) + "/" + str(datual.month) + "/" + str(datual.year) )
        
        lbldescricao1 = Label(jnl5,text="Descrição gastos sem retorno", width=30, font="arial 12 bold", bg="#fff")
        lbldescricao1.place(x=10, y=100)
        lbldescricao2 = Label(jnl5,text="Descrição gastos com retorno", width=30, font="arial 12 bold", bg="#fff")
        lbldescricao2.place(x=415, y=100)
        
        lbldescricao11 = Label(jnl5,text="valor", width=8, font="arial 12 bold", bg="#fff")
        lbldescricao11.place(x=312, y=100)
        lbldescricao22 = Label(jnl5,text="valor", width=8, font="arial 12 bold", bg="#fff")
        lbldescricao22.place(x=687, y=100)
        
        fonte = self.fonteLabel
        
        #coluna gastos sem retorno
        
        egastos11 = Entry(jnl5,width=35, font=fonte)
        egastos11.place(x=10, y=120) 
        egastos111 = Entry(jnl5,width=10, font=fonte)
        egastos111.place(x=310, y=120) 
        
        egastos12 = Entry(jnl5,width=35, font=fonte)
        egastos12.place(x=10, y=143) 
        egastos122 = Entry(jnl5,width=10, font=fonte)
        egastos122.place(x=310, y=143) 
        
        egastos13 = Entry(jnl5,width=35, font=fonte)
        egastos13.place(x=10, y=166) 
        egastos133 = Entry(jnl5,width=10, font=fonte)
        egastos133.place(x=310, y=166) 
        
        egastos14 = Entry(jnl5,width=35, font=fonte)
        egastos14.place(x=10, y=189) 
        egastos144 = Entry(jnl5,width=10, font=fonte)
        egastos144.place(x=310, y=189) 
        
        egastos15 = Entry(jnl5,width=35, font=fonte)
        egastos15.place(x=10, y=212) 
        egastos155 = Entry(jnl5,width=10, font=fonte)
        egastos155.place(x=310, y=212) 
        
        egastos16 = Entry(jnl5,width=35, font=fonte)
        egastos16.place(x=10, y=235) 
        egastos166 = Entry(jnl5,width=10, font=fonte)
        egastos166.place(x=310, y=235) 
        
        egastos17 = Entry(jnl5,width=35, font=fonte)
        egastos17.place(x=10, y=258) 
        egastos177 = Entry(jnl5,width=10, font=fonte)
        egastos177.place(x=310, y=258) 
        
        egastos18 = Entry(jnl5,width=35, font=fonte)
        egastos18.place(x=10, y=281) 
        egastos188 = Entry(jnl5,width=10, font=fonte)
        egastos188.place(x=310, y=281) 
        
        egastos19 = Entry(jnl5,width=35, font=fonte)
        egastos19.place(x=10, y=304) 
        egastos199 = Entry(jnl5,width=10, font=fonte)
        egastos199.place(x=310, y=304) 
        
        egastos10 = Entry(jnl5,width=35, font=fonte)
        egastos10.place(x=10, y=327) 
        egastos100 = Entry(jnl5,width=10, font=fonte)
        egastos100.place(x=310, y=327) 
        
        # coluna gastos com retorno
        
        egastos21 = Entry(jnl5,width=30, font=fonte)
        egastos21.place(x=410, y=120)
        egastos211 = Entry(jnl5,width=10, font=fonte)
        egastos211.place(x=680, y=120) 
        
        egastos22 = Entry(jnl5,width=35, font=fonte)
        egastos22.place(x=410, y=143) 
        egastos222 = Entry(jnl5,width=10, font=fonte)
        egastos222.place(x=680, y=143) 

        egastos23 = Entry(jnl5,width=35, font=fonte)
        egastos23.place(x=410, y=166) 
        egastos233 = Entry(jnl5,width=10, font=fonte)
        egastos233.place(x=680, y=166) 
        
        egastos24 = Entry(jnl5,width=35, font=fonte)
        egastos24.place(x=410, y=189) 
        egastos244 = Entry(jnl5,width=10, font=fonte)
        egastos244.place(x=680, y=189) 
        
        egastos25 = Entry(jnl5,width=35, font=fonte)
        egastos25.place(x=410, y=212) 
        egastos255 = Entry(jnl5,width=10, font=fonte)
        egastos255.place(x=680, y=212) 
        
        egastos26 = Entry(jnl5,width=35, font=fonte)
        egastos26.place(x=410, y=235) 
        egastos266 = Entry(jnl5,width=10, font=fonte)
        egastos266.place(x=680, y=235) 
        
        egastos27 = Entry(jnl5,width=35, font=fonte)
        egastos27.place(x=410, y=281) 
        egastos277 = Entry(jnl5,width=10, font=fonte)
        egastos277.place(x=680, y=281) 
        
        egastos28 = Entry(jnl5,width=35, font=fonte)
        egastos28.place(x=410, y=304) 
        egastos288 = Entry(jnl5,width=10, font=fonte)
        egastos288.place(x=680, y=304) 
        
        egastos29 = Entry(jnl5,width=35, font=fonte)
        egastos29.place(x=410, y=327) 
        egastos299 = Entry(jnl5,width=10, font=fonte)
        egastos299.place(x=680, y=327) 
        
        egastos20 = Entry(jnl5,width=35, font=fonte)
        egastos20.place(x=410, y=258) 
        egastos200 = Entry(jnl5,width=10, font=fonte)
        egastos200.place(x=680, y=258) 
        
        
        
        lbltotalsretorno = Label(jnl5, text="Total:",font= fonte, width=10,anchor=W, bg="white")
        lbltotalsretorno.place(x=200 , y=350)
        lbltotalsretornov = Label(jnl5, text="",font= fonte, width=10, bg="white")
        lbltotalsretornov.place(x=310, y=350)
        lbltotalcretorno = Label(jnl5, text="Total:",font= fonte, width=10,anchor=W ,bg="white")
        lbltotalcretorno.place(x=610 , y=350)
        lbltotalcretornov = Label(jnl5, text="",font= fonte, width=10, anchor=E,  bg="white")
        lbltotalcretornov.place(x=670, y=350)
        
        def calcular():
            w = [egastos11.get(),egastos12.get(),egastos13.get(),egastos14.get(),egastos15.get(),egastos16.get(),egastos17.get(),egastos18.get(),egastos19.get(),egastos10.get()]
            x = [egastos111.get(),egastos122.get(),egastos133.get(),egastos144.get(),egastos155.get(),egastos166.get(),egastos177.get(),egastos188.get(),egastos199.get(),egastos100.get()]
            y = [egastos21.get(),egastos22.get(),egastos23.get(),egastos24.get(),egastos25.get(),egastos26.get(),egastos27.get(),egastos28.get(),egastos29.get(),egastos20.get()]
            z = [egastos211.get(),egastos222.get(),egastos233.get(),egastos244.get(),egastos255.get(),egastos266.get(),egastos277.get(),egastos288.get(),egastos299.get(),egastos200.get()]
            
            descricaogastossr = []
            valoresgastossr = []
            descricaogastoscr = []
            valoresgastoscr = []
            
            ttsrt = 0
            ttcrt = 0
            
            for num in range(0,10):            
                if w[num] != "" and x[num] != "":
                    ttsrt += float(x[num].replace(",","."))
                    descricaogastossr.append(w[num])
                    valoresgastossr.append(x[num])
                    
                else:
                    pass
                if y[num] != "" and z[num] != "":
                    ttcrt += float(z[num].replace(",","."))
                    descricaogastoscr.append(y[num])
                    valoresgastoscr.append(z[num])
                    
                else:
                    pass
                    
                if w[num] != "" and x[num] == "":
                    messagebox.showinfo(title="ERRO", message=f'ITEM "{w[num]}" SEM VALOR DESCRITO')
                else:
                    
                    pass
                if y[num] != "" and z[num] == "":
                    messagebox.showinfo(title="ERRO", message=f'ITEM "{y[num]}" SEM VALOR DESCRITO')
                else:    
                    pass
                
                if w[num] == "" and x[num] != "":
                    messagebox.showinfo(title="ERRO", message=f'ITEM "{x[num]}" SEM DESCRIÇÃO')
                else:
                    pass
                if y[num] == "" and z[num] != "":
                    messagebox.showinfo(title="ERRO", message=f'ITEM "{z[num]}" SEM VALOR DESCRIÇÃO')
                else:    
                    pass
                
            lbltotalsretornov["text"] = format(ttsrt , ".2f")
            lbltotalcretornov["text"] = format(ttcrt , ".2f")
            
            def enviar():
                
                confirmacao = messagebox.askyesno(title="Confirmar", message="Confirmar Envio ?")
                
                if confirmacao:    
                    with open (r"c:\FLPrograms\Read\Config\Texts\Users.txt","r") as users:
                        for usuarios in users:
                            usuario = usuarios
        
                    diaa = edia.get()
                    
                    ts = len(descricaogastossr)
                    if ts > 0:
                        for rows in range(0,ts):
                            dgs = descricaogastossr[rows]
                            vsr = valoresgastossr[rows]
                            self.cursorgastos.execute(f"INSERT INTO gastos (dia , gastosSemRetorno , valorSR,funcionario) VALUES ('{diaa}', '{dgs}', {vsr} , '{usuario}')")
                            self.bankgastos.commit()
                        
                    tc = len(descricaogastoscr)
                    if tc > 0 :
                        for value in range(0,tc):
                            dgc = descricaogastoscr[value]
                            vcr = valoresgastoscr[value]
                            self.cursorgastos.execute(f"INSERT INTO gastos (dia,gastosComRetorno,valorCR, funcionario) VALUES ('{diaa}', '{dgc}' , {vcr} , '{usuario}')")
                            self.bankgastos.commit()
                            
                    jnl5.destroy()
                    self.gerirgastos()  
            
            btnenviar = Button(jnl5,text="Enviar", width=10, bg="#4dc3ff", font=fonte, command=enviar)
            btnenviar.place(x=580, y=390)
        
        btnconfirm = Button(jnl5,text="Confirmar", width=10, bg="#4dc3ff", font=fonte, command=calcular)
        btnconfirm.place(x= 680, y = 390)
        jnl5.iconbitmap(self.icone)  
        jnl5.mainloop()

    def relatorio(self):
        
        jnl = Tk()
        jnl.title("Relatorio")
        jnl.geometry("800x600")
        jnl.state("zoomed")
        nb = ttk.Notebook(jnl)
        nb.place(x=0 , y=0 , width=800 , height=600)
        
        tb1 = Frame(nb)
        tb2 = Frame(nb)
        tb3 = Frame(nb)
        
        nb.add(tb1, text="Inicio" )
        nb.add(tb2, text="Relatorio Mensal")
        nb.add(tb3, text="Relatorio Anual")
        
        self.cursorgastos.execute("SELECT * FROM gastos")
        self.lista = self.cursorgastos.fetchall()
        self.cursorfechamento.execute("SELECT * FROM fechamento")
        self.listafechamento = self.cursorfechamento.fetchall()
        
        #janela de relatorio inicio
        lbltitulo1 = Label(tb1,text="RELATORIO", font=self.fontetitulo)
        lbltitulo1.place(x=250 , y= 30)
        
        tvgastos = ttk.Treeview(tb1, columns=("Dt", "dsg", "Vsr") , show="headings")
        tventrada = ttk.Treeview(tb1, columns=("Dt", "din", "car", "pix") , show="headings")
        tvgastos.place(x=60, y=100)
        tventrada.place(x=400, y=100)
        tvgastos.column("Dt", minwidth=70, width=70)
        tvgastos.column("dsg", minwidth=150, width=150)
        tvgastos.column("Vsr", minwidth=100, width=100)
        
        tventrada.column("Dt", minwidth=70, width=70)
        tventrada.column("din", minwidth=100, width=100)
        tventrada.column("car", minwidth=100, width=100)
        tventrada.column("pix", minwidth=100, width=100)
        
        tvgastos.heading('Dt', text="Data")
        tvgastos.heading('dsg', text="Descrição")
        tvgastos.heading('Vsr', text="Valor")
        tventrada.heading('Dt', text="Data")
        tventrada.heading('din', text="Dinheiro")
        tventrada.heading('car', text="catão")
        tventrada.heading('pix', text="Pix")
        
        diatual = datetime.datetime.now().day
        mesatual = datetime.datetime.now().month
        anoatual = datetime.datetime.now().year
            
        
        
        for row4 in self.lista:
            if row4[1] != None:
                
                dia1 = row4[1][-10:2]
                dia  = int(dia1)
                mes1 = row4[1][-7:5]
                mes = int(mes1)
                anos = row4[1][-4:10]
                ano = int(anos)
        
                if  mes == mesatual and ano == anoatual and (row4[2] != None or row4[4] != None):
                    if row4[2] != None:
                        description = row4[2]
                        money = row4[3]
                        tvgastos.insert("", "end", values=(row4[1],"S.R  " + description,money))
                    else:
                        description = row4[4]
                        money = row4[5]
                        tvgastos.insert("", "end", values=(row4[1],"C.R  " + description,money))
        dias = []
        moneys = []
        cards = []
        pixs = []
        for row5 in self.listafechamento:
            if row5[1] != None:
                ms1 = row5[1][-7:5]
                mes2 = int(ms1)
                ano1 = row5[1][-4:10]
                ano2 = int(ano1)
                if  mes2 == mesatual and ano2 == anoatual and row5[2] != None:
                    
                    if dias.count(row5[1]) == 0: 
                        dias.append(row5[1])
                        moneys.append(row5[3])
                        cards.append(row5[4])
                        pixs.append(row5[6])
                    elif dias.count(row5[1]) == 1:
                        indice = dias.index(row5[1])
                        moneys[indice] += row5[3]
                        cards[indice] += row5[4]
                        pixs[indice] += row5[6]
        for dt in dias:
            idce = dias.index(dt)
            tventrada.insert("", "end", values=(dt, moneys[idce], cards[idce], pixs[idce]))
                    
        #janela 2, relatorio mensal
        lbltitulo2 = Label(tb2,text="RELATORIO MENSAL", font=self.fontetitulo)
        lbltitulo2.place(x=250 , y= 30)
        
        lista_meses = ["Janeiro" , "Fervereiro" , "Março" , "Abril" , "Maio" , "Junho" , "Julho" , "Agosto" , "Setembro" , "Outubro" , "Novembro" , "Dezembro"]
        datual = datetime.datetime.now()
        indice_mes = mesatual - 1
        
        list_mes = ttk.Combobox(tb2, values=lista_meses)
        list_mes.set(lista_meses[indice_mes])
        list_mes.place(x=447 , y=67)
            
        fonte = "arial 12"
        
        lista_anos = []
        for itens in self.lista:
            data = itens[1]
            ano = (data[-4:10])
            num = lista_anos.count(ano)
            if num == 0 :
                lista_anos.append(str(ano))
        
        lista_anos.sort()
        list_ano = ttk.Combobox(tb2, values=lista_anos , width=4)
        list_ano.set(anoatual)
        list_ano.place(x=596 , y=67)
        
        
        tv = ttk.Treeview(tb2, columns=("Dt", "Gsr" , "Vsr") , show="headings")
        tvc = ttk.Treeview(tb2, columns=("Dt",  "Gcr", "Vcr") , show="headings")
        tve = ttk.Treeview(tb2, columns=("data", "dinheiro", "cartao", "aprazo", "pix"), show="headings")
        tv.place(x=10, y=100)
        tvc.place(x=400, y=100)
        tve.place(x=10, y=340)
        
        tv.column("Dt", minwidth=70, width=70)
        tv.column("Gsr", minwidth=200, width=200)
        tv.column("Vsr", minwidth=100, width=100)
        tvc.column("Dt", minwidth=70, width=70)
        tvc.column("Gcr", minwidth=200, width=200)
        tvc.column("Vcr", minwidth=100, width=100)
        tve.column("data", minwidth=70, width=70)
        tve.column("dinheiro", minwidth=100, width=100)
        tve.column("cartao", minwidth=100, width=100)
        tve.column("aprazo", minwidth=100, width=100)
        tve.column("pix", minwidth=100, width=100)
        tv.heading('Dt', text="Data")
        tv.heading('Gsr', text="Gastos Sem Retorno")
        tv.heading('Vsr', text="Valor")
        tvc.heading('Dt', text="Data")
        tvc.heading('Gcr', text="Gastos Com Retorno")
        tvc.heading('Vcr', text="Valor")
        tve.heading("data", text="Data")
        tve.heading("dinheiro", text="Dinheiro")
        tve.heading("cartao", text="Cartão")
        tve.heading("aprazo", text="A Prazo")
        tve.heading("pix", text="Pix/Ticket")
        
        totalsr = 0
        totalcr = 0
        
        for item in self.lista:
            data = item[1]
            if item[2] != None:
                dgastosSR = item[2]
                vgastosSR = item[3] 
                totalsr +=vgastosSR
                
                tv.insert("", "end", values=(data,dgastosSR,vgastosSR))
            
            if item[4] != None:
                    
                dgastosCR = item[4]
                vgastosCR = item[5] 
                totalcr +=vgastosCR
                
                tvc.insert("", "end", values=(data,dgastosCR,vgastosCR))
        
            
            
        def atualizar(): 
        
            tv.destroy()
            tvc.destroy()
            tve.destroy()

            tv2 = ttk.Treeview(tb2, columns=("Dt", "Gsr" , "Vsr") , show="headings")
            tvc2 = ttk.Treeview(tb2, columns=("Dt",  "Gcr", "Vcr") , show="headings")
            tve2 = ttk.Treeview(tb2, columns=("data", "dinheiro", "cartao", "aprazo", "pix"), show="headings")
            tv2.place(x=10, y=100)
            tvc2.place(x=400, y=100)
            tve2.place(x=10, y=340)
            
            tv2.column("Dt", minwidth=0, width=70)
            tv2.column("Gsr", minwidth=0, width=200)
            tv2.column("Vsr", minwidth=0, width=100)
            tvc2.column("Dt", minwidth=0, width=70)
            tvc2.column("Gcr", minwidth=0, width=200)
            tvc2.column("Vcr", minwidth=0, width=100)
            tve2.column("data", minwidth=70, width=70)
            tve2.column("dinheiro", minwidth=100, width=100)
            tve2.column("cartao", minwidth=100, width=100)
            tve2.column("aprazo", minwidth=100, width=100)
            tve2.column("pix", minwidth=100, width=100)    
            tv2.heading('Dt', text="Data")
            tv2.heading('Gsr', text="Gastos Sem Retorno")
            tv2.heading('Vsr', text="Valor")
            tvc2.heading('Dt', text="Data")
            tvc2.heading('Gcr', text="Gastos Com Retorno")
            tvc2.heading('Vcr', text="Valor")
            tve2.heading("data", text="Data")
            tve2.heading("dinheiro", text="Dinheiro")
            tve2.heading("cartao", text="Cartão")
            tve2.heading("aprazo", text="A Prazo")
            tve2.heading("pix", text="Pix/Ticket")
                
            nmes = "0" + str(lista_meses.index(list_mes.get()) + 1) if lista_meses.index(list_mes.get()) < 9 else str(lista_meses.index(list_mes.get()) + 1)
            
            totalsr = 0
            totalcr = 0
            dtsr = []
            dgastsr = []
            vgastsr = []
            dtcr = []
            dgastcr = []
            vgastcr = []
            ttpdia = [[],[],[],[],[],[],[],[],[]]
            for item in self.lista:
                data = item[1]
                dia = (data[:2])
                mes = (data[-7:5])
                ano = (data[-4:10])
                if nmes == mes and list_ano.get() == ano:
                    if item[2] != None:
                        dgastosSR = item[2]
                        vgastosSR = item[3] 
                        totalsr +=vgastosSR
                        dtsr.append(data)
                        dgastsr.append(dgastosSR)
                        vgastsr.append(vgastosSR)
                        if ttpdia[0].count(data) == 0:
                            ttpdia[0].append(data)
                            ttpdia[1].append(vgastosSR)
                        else:
                            ttpdia[1][ttpdia[0].index(data)] += vgastosSR
                        
                        tv2.insert("", "end", values=(data, dgastosSR, vgastosSR))

                    if item[4] != None:
                            
                        dgastosCR = item[4]
                        vgastosCR = item[5] 
                        totalcr +=vgastosCR
                        
                        dtcr.append(data)
                        dgastcr.append(dgastosCR)
                        vgastcr.append(vgastosCR)
                        if ttpdia[2].count(data) == 0:
                            ttpdia[2].append(data)
                            ttpdia[3].append(vgastosSR)
                        else:
                            ttpdia[3][ttpdia[2].index(data)] += vgastosSR
                        
                        tvc2.insert("", "end", values=(data, dgastosCR, vgastosCR))
            
            dent = []
            vdent = []
            vcent = []
            vapent = []
            vpent = []
            for seg in self.listafechamento:
                if seg[1] != None:
                    dte = seg[1]
                    diia = (dte[:2])
                    mees = (dte[-7:5])
                    anoo = (dte[-4:10])
                    
                    if nmes == mees and list_ano.get() == anoo:
                        vdin = seg[3]
                        vcard = seg[4]
                        vap = seg[5]
                        vpix = seg[6]
                        
                        if ttpdia[4].count(dte) == 0:
                            ttpdia[4].append(dte)
                            ttpdia[5].append(vdin)
                            ttpdia[6].append(vcard)
                            ttpdia[7].append(vap)
                            ttpdia[8].append(vpix)
                        else:
                            ttpdia[5][ttpdia[4].index(dte)] += vdin
                            ttpdia[6][ttpdia[4].index(dte)] += vcard
                            ttpdia[7][ttpdia[4].index(dte)] += vap
                            ttpdia[8][ttpdia[4].index(dte)] += vpix
                            
                        dent.append(dte)
                        vdent.append(vdin)
                        vcent.append(vcard)
                        vapent.append(vap)
                        vpent.append(vpix)
                        tve2.insert("", "end", values=(dte, vdin, vcard, vap, vpix))
            
            r_mensal = openpyxl.load_workbook(r"c:\FLPrograms\Read\R22M233.xlsx")
            sh_mensal = r_mensal["Table"]
            ind = 0
            
            for cell in sh_mensal.iter_rows(min_row=6, max_row=46):
                try:
                    if dtsr[ind] == None and dtcr[ind] == None:
                        break
                    cell[1].value = dtsr[ind] if dtsr[ind] != None else ""
                    cell[2].value = dgastsr[ind] if dgastsr[ind] != None else ""
                    cell[3].value = vgastsr[ind] if vgastsr[ind] != None else ""
                    ind += 1
                except:
                    break
            ei = 0
            for cel in sh_mensal.iter_rows(min_row=6, max_row=46):
                try:
                    cel[5].value = dtcr[ei] if dtcr[ei] != None else ""
                    cel[6].value = dgastcr[ei] if dgastcr[ei] != None else ""
                    cel[7].value = vgastcr[ei] if vgastcr[ei] != None else ""
                    ei += 1
                except:
                    break
            eii = 0
            for cl in sh_mensal.iter_rows(min_row=6, max_row=36):
                try:  
                    cl[10].value = ttpdia[4][eii] if ttpdia[4][eii] != None else ""
                    cl[11].value = ttpdia[5][eii] if ttpdia[5][eii] != None else ""
                    cl[12].value = ttpdia[6][eii] if ttpdia[6][eii] != None else ""
                    cl[13].value = ttpdia[7][eii] if ttpdia[7][eii] != None else ""
                    cl[14].value = ttpdia[8][eii] if ttpdia[8][eii] != None else ""
                    eii += 1
                    if len(ttpdia[4]) == eii:
                        break
                except:
                    break
            ee = 0
            dtas = ttpdia[0] + ttpdia[2] + ttpdia[4]
            c = []
            for nun in dtas:
                if c.count(nun) == 0:
                    c.insert(0, nun)
            c.sort()

            for cels in sh_mensal.iter_rows(min_row=6, max_row=36):
                try:
                    cels[19].value = c[ee]
                    ee += 1
                    if len(c) == ee:
                        break
                except:
                    break
            r_mensal.save(f"c:\FLPrograms\Relatorios\Relatorio Mensal {list_mes.get()} {ano}.xlsx")

                
            
            btnimprimir = Button(tb2,text="IMPRIMIR" ,bg="#4dc3ff", font="arial 8", command=lambda:self.imprimir_relatorio(f"Relatorio Mensal {list_mes.get()} {ano}.xlsx"))
            btnimprimir.place(x=707, y=65)
        
        
        btnatualizar = Button(tb2, text="Atualizar", font="arial 8" ,bg="#4dc3ff", command=atualizar)
        btnatualizar.place(x=646 , y=65)
        
        
        #janela 3 relatorio anoal
        
        
        l_ano = []
        for itns in self.lista:
            dat = itns[1]
            year = (dat[-4:10])
            nm = l_ano.count(year)
            if nm == 0 :
                l_ano.append(str(year))
        
        l_ano.sort()
        self.l_anos = ttk.Combobox(tb3, values=lista_anos , width=4)
        self.l_anos.set(anoatual)
        self.l_anos["font"] = "arial 11"
        self.l_anos.place(x=615 , y=67)
        
        def bscar():
            self.ano = self.l_anos.get()
            self.bt_imprimir = Button(tb3,text="", font="arial 9", bg="#4dc3ff", command=lambda:self.imprimir_relatorio(f"Relatorio Anual {self.ano}.xlsx"))
            self.bt_imprimir.place(x=730, y=65)
            self.buscar_gastos()
            
        btnbuscar = Button(tb3, text="Buscar", font="arial 9" ,bg="#4dc3ff", command=bscar)
        btnbuscar.place(x=677 , y=65)
        
        lbltitulo3 = Label(tb3,text="RELATORIO ANUAL", font=self.fontetitulo)
        lbltitulo3.place(x=250 , y= 30)
        
        self.lbljan = Label(tb3, text="Janeiro", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblfev = Label(tb3, text="Fevereiro", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblmar = Label(tb3, text="Março", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblabr = Label(tb3, text="Abril", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblmai = Label(tb3, text="Maio", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbljun = Label(tb3, text="Junho", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbljul = Label(tb3, text="Julho", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblago = Label(tb3, text="Agosto", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblset = Label(tb3, text="Setembro", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblout = Label(tb3, text="Outubro", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblnov = Label(tb3, text="Novembro", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbldez = Label(tb3, text="Dezembro", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblttl = Label(tb3, text="Total", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        
        self.lbljan.place(x=100, y=150)
        self.lblfev.place(x=100, y=175)
        self.lblmar.place(x=100, y=200)
        self.lblabr.place(x=100, y=225)
        self.lblmai.place(x=100, y=250)
        self.lbljun.place(x=100, y=275)
        self.lbljul.place(x=100, y=300)
        self.lblago.place(x=100, y=325)
        self.lblset.place(x=100, y=350)
        self.lblout.place(x=100, y=375)
        self.lblnov.place(x=100, y=400)
        self.lbldez.place(x=100, y=425)
        self.lblttl.place(x=100, y=450)
        
        self.lblsaida = Label(tb3, text="DESPESAS", width=20, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblsaida.place(x=201, y=100)
        self.lbltsaida = Label(tb3, text="DESPESAS S RETORNO", width=20, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbltentrada = Label(tb3, text="TOTAL ENTRADAS", width=20, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblttotal = Label(tb3, text="SALDO FINAL DO ANO", width=20, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbltsaida.place(x=201, y=500)
        self.lbltentrada.place(x=402, y=500)
        self.lblttotal.place(x=603, y=500)
        
        self.lbtsaida = Label(tb3, text="", width=20, bg="#fff", bd=1, relief="solid", font=self.fonteLabel)
        self.lbtentrada = Label(tb3, text="", width=20, bg="#fff", bd=1, relief="solid", font=self.fonteLabel)
        self.lbttotal = Label(tb3, text="", width=20, bg="#fff", bd=1, relief="solid", font=self.fonteLabel)
        self.lbtsaida.place(x=201, y=525)
        self.lbtentrada.place(x=402, y=525)
        self.lbttotal.place(x=603, y=525)
        #coluna sem retorno
        self.lbldessr = Label(tb3, text="S. Retorno", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblvjan = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvfev = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvmar = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvabr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvmai = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvjun = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvjul = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvago = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvset = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvout = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvnov = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblvdez = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblttlsr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        
        self.lbldessr.place(x=201, y=125)
        self.lblvjan.place(x=201, y=150)
        self.lblvfev.place(x=201, y=175)
        self.lblvmar.place(x=201, y=200)
        self.lblvabr.place(x=201, y=225)
        self.lblvmai.place(x=201, y=250)
        self.lblvjun.place(x=201, y=275)
        self.lblvjul.place(x=201, y=300)
        self.lblvago.place(x=201, y=325)
        self.lblvset.place(x=201, y=350)
        self.lblvout.place(x=201, y=375)
        self.lblvnov.place(x=201, y=400)
        self.lblvdez.place(x=201, y=425)
        self.lblttlsr.place(x=201, y=450)
        
        #coluna com retorno
        self.lbldescr = Label(tb3, text="C. Retorno", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lblcjan = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcfev = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcmar = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcabr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcmai = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcjun = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcjul = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcago = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcset = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcout = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcnov = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblcdez = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblttlcr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        
        self.lbldescr.place(x=302, y=125)
        self.lblcjan.place(x=302, y=150)
        self.lblcfev.place(x=302, y=175)
        self.lblcmar.place(x=302, y=200)
        self.lblcabr.place(x=302, y=225)
        self.lblcmai.place(x=302, y=250)
        self.lblcjun.place(x=302, y=275)
        self.lblcjul.place(x=302, y=300)
        self.lblcago.place(x=302, y=325)
        self.lblcset.place(x=302, y=350)
        self.lblcout.place(x=302, y=375)
        self.lblcnov.place(x=302, y=400)
        self.lblcdez.place(x=302, y=425)
        self.lblttlcr.place(x=302, y=450)
        
        #coluna entrada de dinheiro
        self.lblentrada = Label(tb3, text="ENTRADA GERAL", width=40, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbddes = Label(tb3, text="DINHEIRO", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbdjan = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdfev = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdmar = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdabr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdmai = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdjun = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdjul = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdago = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdset = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdout = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbdnov = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbddez = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblttld = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        
        self.lblentrada.place(x=408, y=100)
        self.lbddes.place(x=408, y=125)
        self.lbdjan.place(x=408, y=150)
        self.lbdfev.place(x=408, y=175)
        self.lbdmar.place(x=408, y=200)
        self.lbdabr.place(x=408, y=225)
        self.lbdmai.place(x=408, y=250)
        self.lbdjun.place(x=408, y=275)
        self.lbdjul.place(x=408, y=300)
        self.lbdago.place(x=408, y=325)
        self.lbdset.place(x=408, y=350)
        self.lbdout.place(x=408, y=375)
        self.lbdnov.place(x=408, y=400)
        self.lbddez.place(x=408, y=425)
        self.lblttld.place(x=408, y=450)
        
        #coluna entrada de cartao
        self.lbcdes = Label(tb3, text="CARTÃO", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbcjan = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcfev = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcmar = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcabr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcmai = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcjun = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcjul = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcago = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcset = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcout = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcnov = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbcdez = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblttlc = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        
        self.lbcdes.place(x=509, y=125)
        self.lbcjan.place(x=509, y=150)
        self.lbcfev.place(x=509, y=175)
        self.lbcmar.place(x=509, y=200)
        self.lbcabr.place(x=509, y=225)
        self.lbcmai.place(x=509, y=250)
        self.lbcjun.place(x=509, y=275)
        self.lbcjul.place(x=509, y=300)
        self.lbcago.place(x=509, y=325)
        self.lbcset.place(x=509, y=350)
        self.lbcout.place(x=509, y=375)
        self.lbcnov.place(x=509, y=400)
        self.lbcdez.place(x=509, y=425)
        self.lblttlc.place(x=509, y=450)
        
        #entrada pix e ticket
        self.lbpdes = Label(tb3, text="PIX/TICKET", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbpjan = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpfev = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpmar = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpabr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpmai = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpjun = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpjul = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpago = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpset = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpout = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpnov = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbpdez = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblttlp = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        
        self.lbpdes.place(x=610, y=125)
        self.lbpjan.place(x=610, y=150)
        self.lbpfev.place(x=610, y=175)
        self.lbpmar.place(x=610, y=200)
        self.lbpabr.place(x=610, y=225)
        self.lbpmai.place(x=610, y=250)
        self.lbpjun.place(x=610, y=275)
        self.lbpjul.place(x=610, y=300)
        self.lbpago.place(x=610, y=325)
        self.lbpset.place(x=610, y=350)
        self.lbpout.place(x=610, y=375)
        self.lbpnov.place(x=610, y=400)
        self.lbpdez.place(x=610, y=425)
        self.lblttlp.place(x=610, y=450)
        
        #coluna total entrada
        self.lbtdes = Label(tb3, text="TOTAL", width=10, bg="#ffff00", bd=1, relief="solid", font=self.fonteLabel)
        self.lbtjan = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtfev = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtmar = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtabr = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtmai = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtjun = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtjul = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtago = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtset = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtout = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtnov = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lbtdez = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        self.lblttlt = Label(tb3, text="", width=10, font=self.fonteLabel, anchor=E, bd=1, relief="solid", bg="#fff")
        
        self.lbtdes.place(x=711, y=125)
        self.lbtjan.place(x=711, y=150)
        self.lbtfev.place(x=711, y=175)
        self.lbtmar.place(x=711, y=200)
        self.lbtabr.place(x=711, y=225)
        self.lbtmai.place(x=711, y=250)
        self.lbtjun.place(x=711, y=275)
        self.lbtjul.place(x=711, y=300)
        self.lbtago.place(x=711, y=325)
        self.lbtset.place(x=711, y=350)
        self.lbtout.place(x=711, y=375)
        self.lbtnov.place(x=711, y=400)
        self.lbtdez.place(x=711, y=425)
        self.lblttlt.place(x=711, y=450)
        
        
        self.totalsr3 = 0
        self.totalcr3 = 0
        self.ttpmessr = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.ttpmescr = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        
        jnl.iconbitmap(self.icone)
        jnl.mainloop()

    def buscar_gastos(self):
        
        
        for item in self.lista:
            
            data = item[1]
            dia = (data[:2])
            mes = (data[-7:5])
            ano = (data[-4:10])
            
            if self.l_anos.get() == ano:
                for cdmes in range(0,12):
                    
                    cd_mes = ("0" + str((cdmes + 1))) if (cdmes + 1) < 10 else str((cdmes + 1))
                    if item[2] != None and mes == cd_mes:
                        
                        dgastosSR = item[2]
                        vgastosSR = item[3] 
                        self.totalsr3 +=vgastosSR
                        self.ttpmessr[cdmes] += vgastosSR
                        
                    elif item[4] != None and mes == cd_mes:
                            
                        dgastosCR = item[4]
                        vgastosCR = item[5] 
                        self.totalcr3 +=vgastosCR
                        self.ttpmescr[cdmes] += vgastosCR
                    else:
                        pass
            else:
                pass
        
        self.totaldinheiro = 0
        self.totalcartao = 0
        self.totalpix = 0
        self.totalt = 0
        self.totalpmesd = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totalpmesc = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totalpmesp = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totalpmes = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        
        for itm in self.listafechamento:
            
            if itm[1] != None:
                
                dta2 = itm[1]
                dia2 = (dta2[:2])
                mes2 = (dta2[-7:5])
                ano2 = (dta2[-4:10])
                
                if self.l_anos.get() == ano:
                    for cmes in range(0,12):
                        
                        cd_mes = ("0" + str((cmes + 1))) if (cmes + 1) < 10 else str((cmes + 1))
                        if itm[1] != None and mes2 == cd_mes:
                            
                            dinheiro = itm[3]
                            cartao = itm[4] 
                            pix = itm[5]
                            self.totaldinheiro += dinheiro
                            self.totalpmesd[cmes] += dinheiro
                            self.totalcartao += cartao
                            self.totalpmesc[cmes] += cartao
                            self.totalpix += pix
                            self.totalpmesp[cmes] += pix
                            
                        else:
                            pass
                else:
                    pass
        for i in self.totalpmesd:
            self.totalpmes[self.totalpmesd.index(i)] = i
        for j in self.totalpmesc:
            self.totalpmes[self.totalpmesc.index(j)] += j
        for k in self.totalpmesp:
            self.totalpmes[self.totalpmesp.index(k)] += k
            
        self.totalt = self.totaldinheiro + self.totalcartao + self.totalpix
        self.lbtotalsaldo = (self.totalt - self.totalsr3)            
        
                        
        
        #GASTOS SEM RETORNO
        self.lblvjan["text"] = str(format(self.ttpmessr[0], ".2f")).replace(".", ",") if self.ttpmessr[0] != 0 else "0,00"
        self.lblvfev["text"] = str(format(self.ttpmessr[1], ".2f")).replace(".", ",") if self.ttpmessr[1] != 0 else "0,00"
        self.lblvmar["text"] = str(format(self.ttpmessr[2], ".2f")).replace(".", ",") if self.ttpmessr[2] != 0 else "0,00"
        self.lblvabr["text"] = str(format(self.ttpmessr[3], ".2f")).replace(".", ",") if self.ttpmessr[3] != 0 else "0,00"
        self.lblvmai["text"] = str(format(self.ttpmessr[4], ".2f")).replace(".", ",") if self.ttpmessr[4] != 0 else "0,00"
        self.lblvjun["text"] = str(format(self.ttpmessr[5], ".2f")).replace(".", ",") if self.ttpmessr[5] != 0 else "0,00"
        self.lblvjul["text"] = str(format(self.ttpmessr[6], ".2f")).replace(".", ",") if self.ttpmessr[6] != 0 else "0,00"
        self.lblvago["text"] = str(format(self.ttpmessr[7], ".2f")).replace(".", ",") if self.ttpmessr[7] != 0 else "0,00"
        self.lblvset["text"] = str(format(self.ttpmessr[8], ".2f")).replace(".", ",") if self.ttpmessr[8] != 0 else "0,00"
        self.lblvout["text"] = str(format(self.ttpmessr[9], ".2f")).replace(".", ",") if self.ttpmessr[9] != 0 else "0,00"
        self.lblvnov["text"] = str(format(self.ttpmessr[10], ".2f")).replace(".", ",") if self.ttpmessr[10] != 0 else "0,00"
        self.lblvdez["text"] = str(format(self.ttpmessr[11], ".2f")).replace(".", ",") if self.ttpmessr[11] != 0 else "0,00"
        self.lblttlsr["text"] = str(format(self.totalsr3, ".2f")).replace(".", ",")
        
        #GASTOS COM RETORNO
        self.lblcjan["text"] = str(format(self.ttpmescr[0], ".2f")).replace(".", ",") if self.ttpmescr[0] != 0 else "0,00"
        self.lblcfev["text"] = str(format(self.ttpmescr[1], ".2f")).replace(".", ",") if self.ttpmescr[1] != 0 else "0,00"
        self.lblcmar["text"] = str(format(self.ttpmescr[2], ".2f")).replace(".", ",") if self.ttpmescr[2] != 0 else "0,00"
        self.lblcabr["text"] = str(format(self.ttpmescr[3], ".2f")).replace(".", ",") if self.ttpmescr[3] != 0 else "0,00"
        self.lblcmai["text"] = str(format(self.ttpmescr[4], ".2f")).replace(".", ",") if self.ttpmescr[4] != 0 else "0,00"
        self.lblcjun["text"] = str(format(self.ttpmescr[5], ".2f")).replace(".", ",") if self.ttpmescr[5] != 0 else "0,00"
        self.lblcjul["text"] = str(format(self.ttpmescr[6], ".2f")).replace(".", ",") if self.ttpmescr[6] != 0 else "0,00"
        self.lblcago["text"] = str(format(self.ttpmescr[7], ".2f")).replace(".", ",") if self.ttpmescr[7] != 0 else "0,00"
        self.lblcset["text"] = str(format(self.ttpmescr[8], ".2f")).replace(".", ",") if self.ttpmescr[8] != 0 else "0,00"
        self.lblcout["text"] = str(format(self.ttpmescr[9], ".2f")).replace(".", ",") if self.ttpmescr[9] != 0 else "0,00"
        self.lblcnov["text"] = str(format(self.ttpmescr[10], ".2f")).replace(".", ",") if self.ttpmescr[10] != 0 else "0,00"
        self.lblcdez["text"] = str(format(self.ttpmescr[11], ".2f")).replace(".", ",") if self.ttpmescr[11] != 0 else "0,00"
        self.lblttlcr["text"] = str(format(self.totalcr3, ".2f")).replace(".", ",")
        
        #ENTRADA DINHEIRO
        self.lbdjan["text"] = str(format(self.totalpmesd[0], ".2f")).replace(".", ",") if self.totalpmesd[0] != 0 else "0,00"
        self.lbdfev["text"] = str(format(self.totalpmesd[1], ".2f")).replace(".", ",") if self.totalpmesd[1] != 0 else "0,00"
        self.lbdmar["text"] = str(format(self.totalpmesd[2], ".2f")).replace(".", ",") if self.totalpmesd[2] != 0 else "0,00"
        self.lbdabr["text"] = str(format(self.totalpmesd[3], ".2f")).replace(".", ",") if self.totalpmesd[3] != 0 else "0,00"
        self.lbdmai["text"] = str(format(self.totalpmesd[4], ".2f")).replace(".", ",") if self.totalpmesd[4] != 0 else "0,00"
        self.lbdjun["text"] = str(format(self.totalpmesd[5], ".2f")).replace(".", ",") if self.totalpmesd[5] != 0 else "0,00"
        self.lbdjul["text"] = str(format(self.totalpmesd[6], ".2f")).replace(".", ",") if self.totalpmesd[6] != 0 else "0,00"
        self.lbdago["text"] = str(format(self.totalpmesd[7], ".2f")).replace(".", ",") if self.totalpmesd[7] != 0 else "0,00"
        self.lbdset["text"] = str(format(self.totalpmesd[8], ".2f")).replace(".", ",") if self.totalpmesd[8] != 0 else "0,00"
        self.lbdout["text"] = str(format(self.totalpmesd[9], ".2f")).replace(".", ",") if self.totalpmesd[9] != 0 else "0,00"
        self.lbdnov["text"] = str(format(self.totalpmesd[10], ".2f")).replace(".", ",") if self.totalpmesd[10] != 0 else "0,00"
        self.lbddez["text"] = str(format(self.totalpmesd[11], ".2f")).replace(".", ",") if self.totalpmesd[11] != 0 else "0,00"
        self.lblttld["text"] = str(format(self.totaldinheiro, ".2f")).replace(".", ",")
        
        #ENTRADA CARTAO
        self.lbcjan["text"] = str(format(self.totalpmesc[0], ".2f")).replace(".", ",") if self.totalpmesc[0] != 0 else "0,00"
        self.lbcfev["text"] = str(format(self.totalpmesc[1], ".2f")).replace(".", ",") if self.totalpmesc[1] != 0 else "0,00"
        self.lbcmar["text"] = str(format(self.totalpmesc[2], ".2f")).replace(".", ",") if self.totalpmesc[2] != 0 else "0,00"
        self.lbcabr["text"] = str(format(self.totalpmesc[3], ".2f")).replace(".", ",") if self.totalpmesc[3] != 0 else "0,00"
        self.lbcmai["text"] = str(format(self.totalpmesc[4], ".2f")).replace(".", ",") if self.totalpmesc[4] != 0 else "0,00"
        self.lbcjun["text"] = str(format(self.totalpmesc[5], ".2f")).replace(".", ",") if self.totalpmesc[5] != 0 else "0,00"
        self.lbcjul["text"] = str(format(self.totalpmesc[6], ".2f")).replace(".", ",") if self.totalpmesc[6] != 0 else "0,00"
        self.lbcago["text"] = str(format(self.totalpmesc[7], ".2f")).replace(".", ",") if self.totalpmesc[7] != 0 else "0,00"
        self.lbcset["text"] = str(format(self.totalpmesc[8], ".2f")).replace(".", ",") if self.totalpmesc[8] != 0 else "0,00"
        self.lbcout["text"] = str(format(self.totalpmesc[9], ".2f")).replace(".", ",") if self.totalpmesc[9] != 0 else "0,00"
        self.lbcnov["text"] = str(format(self.totalpmesc[10], ".2f")).replace(".", ",") if self.totalpmesc[10] != 0 else "0,00"
        self.lbcdez["text"] = str(format(self.totalpmesc[11], ".2f")).replace(".", ",") if self.totalpmesc[11] != 0 else "0,00"
        self.lblttlc["text"] = str(format(self.totalcartao, ".2f")).replace(".", ",")
        
        #entrada pix e ticket
        self.lbpjan["text"] = str(format(self.totalpmesp[0], ".2f")).replace(".", ",") if self.totalpmesp[0] != 0 else "0,00"
        self.lbpfev["text"] = str(format(self.totalpmesp[1], ".2f")).replace(".", ",") if self.totalpmesp[1] != 0 else "0,00"
        self.lbpmar["text"] = str(format(self.totalpmesp[2], ".2f")).replace(".", ",") if self.totalpmesp[2] != 0 else "0,00"
        self.lbpabr["text"] = str(format(self.totalpmesp[3], ".2f")).replace(".", ",") if self.totalpmesp[3] != 0 else "0,00"
        self.lbpmai["text"] = str(format(self.totalpmesp[4], ".2f")).replace(".", ",") if self.totalpmesp[4] != 0 else "0,00"
        self.lbpjun["text"] = str(format(self.totalpmesp[5], ".2f")).replace(".", ",") if self.totalpmesp[5] != 0 else "0,00"
        self.lbpjul["text"] = str(format(self.totalpmesp[6], ".2f")).replace(".", ",") if self.totalpmesp[6] != 0 else "0,00"
        self.lbpago["text"] = str(format(self.totalpmesp[7], ".2f")).replace(".", ",") if self.totalpmesp[7] != 0 else "0,00"
        self.lbpset["text"] = str(format(self.totalpmesp[8], ".2f")).replace(".", ",") if self.totalpmesp[8] != 0 else "0,00"
        self.lbpout["text"] = str(format(self.totalpmesp[9], ".2f")).replace(".", ",") if self.totalpmesp[9] != 0 else "0,00"
        self.lbpnov["text"] = str(format(self.totalpmesp[10], ".2f")).replace(".", ",") if self.totalpmesp[10] != 0 else "0,00"
        self.lbpdez["text"] = str(format(self.totalpmesp[11], ".2f")).replace(".", ",") if self.totalpmesp[11] != 0 else "0,00"
        self.lblttlp["text"] = str(format(self.totalpix, ".2f")).replace(".", ",")
        
        #entrada total
        self.lbtjan["text"] = str(format(self.totalpmes[0], ".2f")).replace(".", ",") if self.totalpmes[0] != 0 else "0,00"
        self.lbtfev["text"] = str(format(self.totalpmes[1], ".2f")).replace(".", ",") if self.totalpmes[1] != 0 else "0,00"
        self.lbtmar["text"] = str(format(self.totalpmes[2], ".2f")).replace(".", ",") if self.totalpmes[2] != 0 else "0,00"
        self.lbtabr["text"] = str(format(self.totalpmes[3], ".2f")).replace(".", ",") if self.totalpmes[3] != 0 else "0,00"
        self.lbtmai["text"] = str(format(self.totalpmes[4], ".2f")).replace(".", ",") if self.totalpmes[4] != 0 else "0,00"
        self.lbtjun["text"] = str(format(self.totalpmes[5], ".2f")).replace(".", ",") if self.totalpmes[5] != 0 else "0,00"
        self.lbtjul["text"] = str(format(self.totalpmes[6], ".2f")).replace(".", ",") if self.totalpmes[6] != 0 else "0,00"
        self.lbtago["text"] = str(format(self.totalpmes[7], ".2f")).replace(".", ",") if self.totalpmes[7] != 0 else "0,00"
        self.lbtset["text"] = str(format(self.totalpmes[8], ".2f")).replace(".", ",") if self.totalpmes[8] != 0 else "0,00"
        self.lbtout["text"] = str(format(self.totalpmes[9], ".2f")).replace(".", ",") if self.totalpmes[9] != 0 else "0,00"
        self.lbtnov["text"] = str(format(self.totalpmes[10], ".2f")).replace(".", ",") if self.totalpmes[10] != 0 else "0,00"
        self.lbtdez["text"] = str(format(self.totalpmes[11], ".2f")).replace(".", ",") if self.totalpmes[11] != 0 else "0,00"
        self.lblttlt["text"] = str(format(self.totalt, ".2f")).replace(".", ",")
        
        self.lbtsaida["text"] = str(format(self.totalsr3, ".2f")).replace(".", ",")
        self.lbtentrada["text"] = str(format(self.totalt, ".2f")).replace(".", ",")
        self.lbttotal["text"] = str(format(self.lbtotalsaldo, ".2f")).replace(".", ",")
        if self.lbtotalsaldo < 0:
            self.lbttotal["fg"] = "red"
        elif self.lbtotalsaldo > 0:
            self.lbttotal["fg"] = "green"
        
        r_anual = openpyxl.load_workbook(r"c:\FLPrograms\Read\R22A233.xlsx")
        sh_anual = r_anual["Plan1"]
        II = 0
        for I in sh_anual.iter_rows(min_row=6, max_row=17):
            
            I[2].value = self.ttpmessr[II]
            I[3].value = self.ttpmescr[II]
            I[4].value = self.totalpmesd[II]
            I[5].value = self.totalpmesc[II]
            I[6].value = self.totalpmesp[II]
            II += 1
        r_anual.save(f"c:\FLPrograms\Relatorios\Relatorio Anual {ano}.xlsx")
        self.bt_imprimir["text"]="Imprimir" 
        self.bt_imprimir["bg"]="#00ff00"
        self.lbtotalsaldo = 0
        self.totalsr3 = 0
        self.totalcr3 = 0
        self.ttpmessr = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.ttpmescr = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totaldinheiro = 0
        self.totalcartao = 0
        self.totalpix = 0
        self.totalpmesd = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totalpmesc = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totalpmesp = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totalpmes = [0,0,0,0,0,0,0,0,0,0,0,0,0]
        self.totalt = 0
    
    def imprimir_relatorio(self,arquivo):
        ask = Tk()
        ask.title("IMPRIMIR")
        ask.geometry("350x100")
        ask.resizable(0,0)
        ask["bg"] = "#fff"
        cb_print = ttk.Combobox(ask,values=self.lista_impressoras, font="arial 12",width=20)
        cb_print.place(x=20, y=30)
        cb_print.set(self.impressoraPadrao)
        def printer():
            win32print.SetDefaultPrinter(cb_print.get())
            caminho = r"c:\FLPrograms\Relatorios"
            ShellExecute(0, "print" , arquivo ,  None , caminho , 0  )
            win32print.SetDefaultPrinter(self.impressoraPadrao)
            ask.destroy()
        btn_imprimir = Button(ask, text="IMPRIMIR", font="arial 12", width=10, command=printer)
        btn_imprimir.place(x=100, y= 70)
        btn_cancelar = Button(ask, text="CANCELAR", font="arial 12", width=10, command=ask.destroy)
        btn_cancelar.place(x=200, y=70)
        ask.mainloop()
start = aplicativo()
